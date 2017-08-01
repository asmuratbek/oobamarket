import random

from django.core.management.base import BaseCommand, CommandError
from slugify import slugify
from django.db.utils import IntegrityError
from apps.category.models import Category
from apps.global_category.models import GlobalCategory
from django.conf import settings
import openpyxl as px
import string

from apps.properties.models import Properties, Values


class Command(BaseCommand):
    help = 'Creates model instances from xlsx'

    def handle(self, *args, **options):
        W = px.load_workbook(settings.DUMP_ROOT + '/ooba-test.xlsx')
        wiki = dict()
        first_level = list()
        second_level = list()
        third_level = list()
        second_level_props = list()
        third_level_props = list()
        values = list()
        for index, worksheet in enumerate(W.worksheets[1:]):
            if index == 0:
                for row in worksheet.iter_rows():
                    for index, field in enumerate(row):
                        if field.value:
                            if index == 0:
                                key = field.value
                                wiki[key] = []
                            else:
                                wiki[key].append(str(field.value).capitalize())

            else:
                first_level = list()
                second_level = list()
                third_level = list()
                second_level_props = list()
                third_level_props = list()
                values = list()
                slug = slugify(worksheet.title)
                section, created = GlobalCategory.objects.get_or_create(title=worksheet.title, slug=slug)
                for index, col in enumerate(worksheet.iter_cols(max_col=3)):
                    if index == 0:
                        for index, field in enumerate(col):
                            if field.value:
                                first_level.append([str(field.internal_value).split('.')[:-1], str(field.internal_value).split('.')[-1][1:], worksheet.title])

                    elif index == 1:
                        for index, field in enumerate(col):
                            if field.value and not str(field.value).startswith('('):
                                if field.font.bold:
                                    ids = str(field.internal_value).split('.')[:-1]
                                    category_text = str(field.internal_value).split('.')[-1][2:].strip()
                                    category_text = category_text.capitalize()
                                    category_level = str(field.internal_value).split('.')[-1][:2] if \
                                        str(field.internal_value).split('.')[-1][:2].isdigit() else \
                                        str(field.internal_value).split('.')[-1][0]
                                    if category_level:
                                        ids.append(category_level)
                                    second_level.append([ids, category_text, worksheet.title])
                                else:
                                    look_for_vals = True
                                    property_text = str(field.internal_value)
                                    if property_text.startswith('**'):
                                        property_text = property_text[3:]
                                    elif property_text.startswith(('*', "#", "!")):
                                        if property_text.startswith("#"):
                                            property_text = property_text[2:]
                                            property_text = property_text.capitalize()
                                            look_for_values = False
                                            for column in string.ascii_uppercase[
                                                          string.ascii_uppercase.find(field.column) + 1:]:
                                                if worksheet[column + str(field.row)].value:
                                                    second_level_props.append([ids, property_text, str(worksheet[column + str(field.row)].value), worksheet.title])
                                    else:
                                        property_text = property_text.capitalize()
                                        second_level_props.append([ids, property_text, worksheet.title])
                                    if look_for_vals:
                                        for column in string.ascii_uppercase[string.ascii_uppercase.find(field.column) + 1:]:
                                            if worksheet[column + str(field.row)].value:
                                                if str(worksheet[column + str(field.row)].value).startswith('Wiki'):
                                                    key = str(worksheet[column + str(field.row)].value)[7:]
                                                    if key in wiki.keys():
                                                        values.append([ids, property_text, wiki[key], worksheet.title])
                                                else:
                                                    if not str(worksheet[column + str(field.row)].value).startswith("("):
                                                        values.append([ids, property_text, str(worksheet[column + str(field.row)].value).capitalize(), worksheet.title])

                    elif index == 2:
                        for index, field in enumerate(col):
                            if field.value and not str(field.internal_value).startswith('('):
                                if field.font.bold:
                                    ids = str(field.internal_value).split('.')[:-1]
                                    category_text = str(field.internal_value).split('.')[-1][2:].strip()
                                    category_text = category_text.capitalize()
                                    category_level = str(field.internal_value).split('.')[-1][:2] if \
                                    str(field.internal_value).split('.')[-1][:2].isdigit() else \
                                    str(field.internal_value).split('.')[-1][0]
                                    if category_level:
                                        ids.append(category_level)
                                    third_level.append([ids, category_text, worksheet.title])
                                else:
                                    look_for_values = True
                                    property_text = str(field.internal_value)
                                    if property_text.startswith('**'):
                                        property_text = property_text[3:]
                                    elif property_text.startswith(('*', '#', '!')):
                                        if property_text.startswith("#"):
                                            property_text = property_text[2:]
                                            property_text = property_text.capitalize()
                                            look_for_values = False
                                            for column in string.ascii_uppercase[
                                                          string.ascii_uppercase.find(field.column) + 1:]:
                                                if worksheet[column + str(field.row)].value:
                                                    third_level_props.append([ids, property_text, str(worksheet[column + str(field.row)].value), worksheet.title])
                                    else:
                                        property_text = property_text.capitalize()
                                        third_level_props.append([ids, property_text, worksheet.title])
                                    if look_for_values:
                                        for column in string.ascii_uppercase[string.ascii_uppercase.find(field.column) + 1:]:
                                            if worksheet[column + str(field.row)].value:
                                                if str(worksheet[column + str(field.row)].value).startswith('Wiki'):
                                                    key = str(worksheet[column + str(field.row)].value)[7:]
                                                    if key in wiki.keys():
                                                        values.append([ids, property_text, wiki[key], worksheet.title])
                                                else:
                                                    if not str(worksheet[column + str(field.row)].value).startswith("("):
                                                        values.append([ids, property_text, str(worksheet[column + str(field.row)].value).capitalize(), worksheet.title])

            for category in first_level:
                title = category[1]
                extra_slug = ".".join(category[0])
                slug = slugify(title) + "-" + slugify(section.title) + "-" + extra_slug
                cat, created = Category.objects.get_or_create(title=title, slug=slug, section=section)
                if created:
                    self.stdout.write(self.style.SUCCESS('Создана категория 1-ого уровня "%s"' % title))
                else:
                    self.stdout.write(self.style.ERROR('{} уже существует'.format(title)))

                for sec_lvl_cat in second_level:
                    if category[0][0] == sec_lvl_cat[0][0] and category[-1] == sec_lvl_cat[-1]:
                        extra_slug = '.'.join(sec_lvl_cat[0])
                        title = sec_lvl_cat[1]
                        slug = slugify(title) + "-" + slugify(section.title) + "-" + extra_slug
                        sec_cat, sec_cat_created = Category.objects.get_or_create(title=title, parent=cat, slug=slug, section=section)
                        if sec_cat_created:
                            self.stdout.write(self.style.SUCCESS('Создана категория 2-ого уровня "%s"' % title))
                        else:
                            self.stdout.write(self.style.ERROR('{} уже существует'.format(title)))

                        for prop in second_level_props:
                            if prop[0] == sec_lvl_cat[0] and prop[-1] == sec_lvl_cat[-1]:
                                find_sec_vals = True
                                title = prop[1]
                                if title.startswith('Wiki'):
                                    title = title[7:]
                                extra_slug = ".".join(sec_lvl_cat[0]) + str(sec_lvl_cat[-1])
                                slug = slugify(title) + "-" + slugify(section.title) + "-" + extra_slug
                                if len(prop) == 4:
                                    print(prop)
                                    find_sec_vals = False
                                    sec_prop, sec_prop_created = Properties.objects.get_or_create(title=title,
                                                                                                  slug=slug)
                                    sec_prop.help_text = str(prop[2])
                                    sec_prop.save()
                                else:
                                    print(prop)
                                    sec_prop, sec_prop_created = Properties.objects.get_or_create(title=title, slug=slug)
                                sec_prop.category.add(sec_cat)
                                if sec_prop_created:
                                    self.stdout.write(self.style.SUCCESS('Создано свойство "%s"' % title))
                                else:
                                    self.stdout.write(self.style.SUCCESS('Изменено свойство "%s"' % title))

                                if find_sec_vals:
                                    for val in values:
                                        if prop[0] == val[0] and prop[-1] and val[-1]:
                                            value = val[2]
                                            if sec_prop.title == val[1]:
                                                if isinstance(value, list):
                                                    for v in value:
                                                        created_value, created = Values.objects.get_or_create(value=v,
                                                                                                              properties=sec_prop)
                                                        if created:
                                                            self.stdout.write(
                                                                self.style.SUCCESS('Создано значение "%s"' % v))
                                                        else:
                                                            self.stdout.write(
                                                                self.style.SUCCESS('Изменено значение "%s"' % v))
                                                else:
                                                    created_value, created = Values.objects.get_or_create(value=value, properties=sec_prop)
                                                    if created:
                                                        self.stdout.write(self.style.SUCCESS('Создано значение "%s"' % value))
                                                    else:
                                                        self.stdout.write(self.style.SUCCESS('Изменено значение "%s"' % value))

                        for thrd_lvl_cat in third_level:
                            if sec_lvl_cat[0] == thrd_lvl_cat[0][:-1] and sec_lvl_cat[-1] == thrd_lvl_cat[-1]:
                                extra_slug = ".".join(thrd_lvl_cat[0])
                                title = thrd_lvl_cat[1]
                                slug = slugify(title) + "-" + slugify(section.title) + "-" + extra_slug
                                thrd_cat, thrd_cat_created = Category.objects.get_or_create(title=title, slug=slug, parent=sec_cat, section=section)
                                if thrd_cat_created:
                                    self.stdout.write(self.style.SUCCESS('Создана категория 3-ого уровня "%s"' % title))
                                else:
                                    self.stdout.write(self.style.ERROR('{} уже существует'.format(title)))

                                for prop in third_level_props:
                                    if prop[0] == thrd_lvl_cat[0] and prop[-1] == thrd_lvl_cat[-1]:
                                        find_thrd_vals = True
                                        extra_slug = ".".join(thrd_lvl_cat[0]) + str(thrd_lvl_cat[-1])
                                        title = prop[1]
                                        if title.startswith('Wiki'):
                                            title = title[7:]
                                        slug = slugify(title) + "-" + extra_slug
                                        if len(prop) == 4:
                                            print(prop)
                                            find_thrd_vals = False
                                            thrd_prop, created = Properties.objects.get_or_create(title=title,
                                                                                                 slug=slug)
                                            thrd_prop.help_text = str(prop[2])
                                            thrd_prop.save()
                                        else:
                                            thrd_prop, created = Properties.objects.get_or_create(title=title,
                                                                                                  slug=slug)
                                        thrd_prop.category.add(thrd_cat)
                                        if created:
                                            self.stdout.write(
                                                self.style.SUCCESS('Создано свойство "%s"' % title))
                                        else:
                                            self.stdout.write(
                                                self.style.SUCCESS('Изменено свойство "%s"' % title))
                                        if find_thrd_vals:
                                            for val in values:
                                                if prop[0] == val[0] and prop[-1] == val[-1]:
                                                    value = val[2]
                                                    if thrd_prop.title == val[1]:
                                                        if isinstance(value, list):
                                                            for v in value:
                                                                created_value, created = Values.objects.get_or_create(
                                                                    value=v, properties=thrd_prop)
                                                                if created:
                                                                    self.stdout.write(
                                                                        self.style.SUCCESS('Создано значение "%s"' % v))
                                                                else:
                                                                    self.stdout.write(
                                                                        self.style.SUCCESS('Изменено значение "%s"' % v))
                                                        else:
                                                            created_value, created = Values.objects.get_or_create(value=value, properties=thrd_prop)
                                                            if created:
                                                                self.stdout.write(
                                                                    self.style.SUCCESS('Создано значение "%s"' % value))
                                                            else:
                                                                self.stdout.write(
                                                                    self.style.SUCCESS('Изменено значение "%s"' % value))
